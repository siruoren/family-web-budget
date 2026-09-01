"""
Excel 解析服务 - 读取历史 Excel 工作簿并按结构化为 (条目, 数值) 序列

支持的工作表:
  - 君军之家2021~2025年 / 君军之家年度账单2026年始~  (年度账单)
  - 家庭结余 (月度账户结余快照)

每个年度表的表头结构略有不同, 本解析器以 "列名匹配" 而非 "固定列号"
作为容错策略: 扫描前 4 行表头, 收集 (列号 -> 中文列名), 再按字典匹配到
对应 Item / Account。

家庭结余的结构解析:
  - 大类 (银行理财 / 股市理财 / 流动资金（李）/ 流动资金（朱）/ 外债)
    位于合并单元格行 (Excel row 4), 解析时对每一列 "前向填充" 该大类。
  - 小项 (兴业 / 招行 / ... / 支付宝小荷包 / 现金) 位于下一行。
  - 已知 Excel 合并单元格有两处归属错位, 用 GROUP_OVERRIDE 修正:
        支付宝1 -> 流动资金（李）   (Excel 误置于 股市理财)
        支付宝2 -> 流动资金（朱）   (Excel 误置于 流动资金（李）)
"""
from __future__ import annotations
from dataclasses import dataclass, field
from datetime import datetime, date
from typing import Iterable
import openpyxl


# -------------------------------------------------------------- 条目映射
# Excel 表头列名 -> 默认条目 (category, owner, sub_category, name)
ITEM_HEADER_MAP = {
    "君公司": ("收入", "君", "君公司", "君公司工资"),
    "君其他": ("收入", "君", "君其他", "君其他收入"),
    "军公司": ("收入", "军", "军公司", "军公司工资"),
    "军其他": ("收入", "军", "军其他", "军其他收入"),
    "公积金提取": ("收入", "家庭", "公积金提取", "公积金提取"),
    "理财": ("收入", "家庭", "理财", "理财收益"),
    "其他": ("收入", "家庭", "其他", "其他收入"),
    "公积金贷款": ("支出", "家庭", "公积金贷款", "公积金贷款"),
    "物业费": ("支出", "家庭", "物业费", "物业费"),
    "商贷": ("支出", "家庭", "商贷", "商贷"),
    "车贷": ("支出", "家庭", "车贷", "车贷"),
    "房租": ("支出", "家庭", "房租", "房租"),
    "水费": ("支出", "家庭", "水费", "水费"),
    "电费": ("支出", "家庭", "电费", "电费"),
    "煤气": ("支出", "家庭", "燃气", "燃气"),
    "燃气": ("支出", "家庭", "燃气", "燃气"),
}

# 结余表的账户名 -> 账户 (type, owner, name)
ACCOUNT_HEADER_MAP = {
    "兴业": ("银行理财", "李", "兴业"),
    "招行": ("银行理财", "李", "招行"),
    "农商1": ("银行理财", "李", "农商1"),
    "建设银行": ("银行理财", "李", "建设银行"),
    "中行1": ("银行理财", "李", "中行1"),
    "工商1": ("银行理财", "李", "工商1"),
    "华宝（李）": ("股市理财", "李", "华宝（李）"),
    "华宝（朱）": ("股市理财", "朱", "华宝（朱）"),
    "支付宝1": ("流动资金", "李", "支付宝1"),
    "中行2": ("流动资金", "李", "中行2"),
    "工商2": ("流动资金", "李", "工商2"),
    "农商2": ("流动资金", "李", "农商2"),
    "微信": ("流动资金", "李", "微信"),  # 李的微信
    "支付宝2": ("流动资金", "朱", "支付宝2"),
    "工商银行": ("流动资金", "朱", "工商银行"),
    "其他银行": ("流动资金", "朱", "其他银行"),
    "支付宝": ("流动资金", "朱", "支付宝"),
    "支付宝小荷包": ("流动资金", "朱", "支付宝小荷包"),
    "现金": ("流动资金", "朱", "现金"),
}

# 大类 (合并单元格文本) -> 归属: type, 默认 owner
# 用于从家庭结余的合并单元格推断账户类型与归属
GROUP_TO_TYPE = {
    "银行理财": "银行理财",
    "股市理财": "股市理财",
    "流动资金（李）": "流动资金",
    "流动资金(李)": "流动资金",
    "流动资金（朱）": "流动资金",
    "流动资金(朱)": "流动资金",
    "外债": "外债",
}
GROUP_TO_OWNER = {
    "流动资金（李）": "李", "流动资金(李)": "李",
    "流动资金（朱）": "朱", "流动资金(朱)": "朱",
}

# Excel 合并单元格归属错位修正: 小项名 -> 应属大类
GROUP_OVERRIDE = {
    "支付宝1": "流动资金（李）",  # Excel 把支付宝1 并入 股市理财
    "支付宝2": "流动资金（朱）",  # Excel 把支付宝2 并入 流动资金（李）
}

# 应忽略的小项列 (非账户: 合计/差额/特殊说明/期初余额 等)
_IGNORED_SUBHEADS = {"合计", "差额", "特殊说明", "期初余额", "项目/时间",
                     "时间", "日期", "账单周期", "月初盘点"}


# -------------------------------------------------------------- 工作表清单
def _infer_sheet_kind(name: str) -> str:
    """按工作表名推断类型: entries(年度账单) / balances(结余) / other"""
    if "结余" in name or "盘点" in name:
        return "balances"
    if "年度账单" in name or name.startswith("君军之家") or "李君" in name:
        return "entries"
    return "other"


@dataclass
class SheetInfo:
    name: str
    kind: str  # entries / balances / other
    order: int


def parse_sheet_inventory(path: str) -> list[SheetInfo]:
    """枚举 Excel 全部工作表, 返回 (name, kind, order) 清单

    用于首次导入时把每个 sheet 页登记为左侧大菜单。
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    infos = [
        SheetInfo(name=sname, kind=_infer_sheet_kind(sname), order=i)
        for i, sname in enumerate(wb.sheetnames)
    ]
    wb.close()
    return infos


# -------------------------------------------------------------- 数据结构
@dataclass
class ParsedEntry:
    year: int
    month: int
    item_key: tuple  # (category, owner, sub, name) 用于匹配 Item
    value: float
    note: str = ""
    group: str = ""   # 大类: 收入 / 支出 (即 item_key[0])
    sheet: str = ""   # 来源工作表


@dataclass
class ParsedBalance:
    year: int
    month: int
    account_key: tuple  # (type, owner, name)
    value: float
    note: str = ""
    group: str = ""    # 大类: 银行理财 / 股市理财 / 流动资金（李）/ 流动资金（朱）/ 外债
    sheet: str = ""    # 来源工作表 (家庭结余)


@dataclass
class ParseResult:
    sheet_name: str
    kind: str  # "entries" | "balances"
    entries: list = field(default_factory=list)
    balances: list = field(default_factory=list)
    year_from: int | None = None
    year_to: int | None = None
    rows_scanned: int = 0


# -------------------------------------------------------------- 工具
def _to_float(v) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("元", "")
    if not s or s in {"-", "--", "无", "未知"}:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _to_year_month(d) -> tuple[int, int] | None:
    """从单元格 (datetime / 数字 / 字符串) 提取 (year, month)"""
    if isinstance(d, (datetime, date)):
        return d.year, d.month
    if isinstance(d, (int, float)):
        # Excel 序列号 (44xxx) 或 8 位数字 (20260501)
        if 30000 < d < 80000:
            try:
                dt = openpyxl.utils.datetime.from_excel(d)
                return dt.year, dt.month
            except Exception:
                return None
        s = str(int(d))
        if len(s) == 8:
            return int(s[:4]), int(s[4:6])
    if isinstance(d, str):
        s = d.strip()
        if len(s) == 8 and s.isdigit():
            return int(s[:4]), int(s[4:6])
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日"):
            try:
                dt = datetime.strptime(s, fmt)
                return dt.year, dt.month
            except ValueError:
                continue
    return None


def _scan_headers(rows, max_header_rows: int = 5) -> dict:
    """扫描前 N 行, 合并非空单元格 -> {col_index: cell_value}"""
    headers: dict[int, str] = {}
    for row in rows[:max_header_rows]:
        for idx, v in enumerate(row):
            if v is None or v == "":
                continue
            text = str(v).strip()
            # 跳过合说明性文本 (太长 / 不像列名)
            if len(text) > 24:
                continue
            if idx not in headers:
                headers[idx] = text
    return headers


def _all_header_texts(rows, max_header_rows: int = 6) -> dict[int, list[str]]:
    """收集每列前 N 行所有非空文本 (用于跨行表头匹配)"""
    out: dict[int, list[str]] = {}
    for row in rows[:max_header_rows]:
        for idx, v in enumerate(row):
            if v is None or v == "":
                continue
            text = str(v).strip()
            if len(text) > 24:
                continue
            out.setdefault(idx, []).append(text)
    return out


def _find_date_col(rows, header_texts: dict[int, list[str]],
                   candidates: list[str]) -> int | None:
    """
    智能定位日期列:
      1) 在所有表头行的所有候选列名中匹配
      2) 若无匹配, 扫描数据行找含 datetime 的列
    """
    for col_idx, texts in header_texts.items():
        for t in texts:
            if t in candidates:
                return col_idx
    # 回退: 扫描前 15 个数据行, 找 datetime/序列号最集中的列
    best_col = None
    best_hits = 0
    for col_idx in range(min(10, max((len(r) for r in rows[:20]), default=1))):
        hits = 0
        for row in rows[3:18]:
            if col_idx < len(row) and _to_year_month(row[col_idx]) is not None:
                hits += 1
        if hits > best_hits:
            best_hits = hits
            best_col = col_idx
    return best_col if best_hits >= 2 else None


# -------------------------------------------------------------- 年度表解析
def parse_year_sheet(ws, sheet_name: str) -> ParseResult:
    rows = list(ws.iter_rows(values_only=True))
    header_texts = _all_header_texts(rows, max_header_rows=6)
    headers = _scan_headers(rows, max_header_rows=5)
    col_to_idx = {v: k for k, v in headers.items()}

    date_col = _find_date_col(rows, header_texts,
                              ["发生日期", "日期", "账单周期", "项目/时间"])
    if date_col is None:
        return ParseResult(sheet_name=sheet_name, kind="entries")

    # 收集 (列号 -> 条目 key) - 跨行匹配 (任一表头行命中即算)
    entry_cols: dict[int, tuple] = {}
    for col_idx, texts in header_texts.items():
        for t in texts:
            if t in ITEM_HEADER_MAP:
                entry_cols[col_idx] = ITEM_HEADER_MAP[t]
                break

    result = ParseResult(sheet_name=sheet_name, kind="entries")
    years_seen = set()

    for row in rows[3:]:
        if not row or len(row) <= date_col:
            continue
        ym = _to_year_month(row[date_col])
        if ym is None:
            continue
        year, month = ym
        years_seen.add(year)

        for col_idx, item_key in entry_cols.items():
            if col_idx >= len(row):
                continue
            val = _to_float(row[col_idx])
            if val is None:
                continue
            result.entries.append(ParsedEntry(
                year=year, month=month, item_key=item_key, value=val,
                group=item_key[0], sheet=sheet_name,
            ))
            result.rows_scanned += 1

    if years_seen:
        result.year_from = result.year_to = min(years_seen)
        if len(years_seen) > 1:
            result.year_to = max(years_seen)
    return result


# -------------------------------------------------------------- 结余表解析
def _account_from_group(group: str, subhead: str) -> tuple | None:
    """由 (大类, 小项名) 推导 (type, owner, name); 无法归类返回 None

    owner 始终以大类为准 (流动资金（朱）下的"微信"归属朱),
    避免硬编码映射把同名小项错配到另一归属。
    """
    if not group:
        return None
    type_ = GROUP_TO_TYPE.get(group)
    if not type_:
        return None
    # owner
    if group in GROUP_TO_OWNER:
        owner = GROUP_TO_OWNER[group]
    elif group == "银行理财":
        owner = "李"
    elif group == "股市理财":
        owner = "李" if ("（李）" in subhead or "(李)" in subhead) else "朱"
    elif group == "外债":
        owner = "家庭"
    else:
        owner = "家庭"
    return (type_, owner, subhead)


def _resolve_subhead(group: str, subhead: str) -> tuple | None:
    """解析小项 -> (type, owner, name); 优先按大类推导, 大类缺失时回退硬编码映射"""
    acc = _account_from_group(group, subhead)
    if acc:
        return acc
    # 回退: 大类未知时用已知映射 (兼容结构不规范的工作表)
    return ACCOUNT_HEADER_MAP.get(subhead)


def _find_subhead_row(rows) -> tuple[int | None, int | None]:
    """定位小项行 (含 项目/时间 / 时间 等日期表头) 与日期列号

    注意: "月初盘点" 是标题行 (非小项行), 不作为候选, 避免误判。
    """
    candidates = {"项目/时间", "时间", "日期", "账单周期"}
    for ri, row in enumerate(rows[:8]):
        for ci, v in enumerate(row):
            if v and str(v).strip() in candidates:
                return ri, ci
    # 回退: 第一行含多个已知账户名
    for ri, row in enumerate(rows[:8]):
        hits = sum(1 for v in row if v and str(v).strip() in ACCOUNT_HEADER_MAP)
        if hits >= 3:
            return ri, 0
    return None, None


# -------------------------------------------------------------- 结余表结构提取
@dataclass
class BalanceSubhead:
    """结余表的小项结构 (不依赖是否有数据)"""
    group: str     # 大类
    type: str      # 账户类型
    owner: str     # 归属 (李/朱/家庭)
    name: str      # 小项名


def extract_balance_structure(ws, sheet_name: str) -> list[BalanceSubhead]:
    """从家庭结余表头提取全部 大类->小项 结构 (即便某些小项无数据也登记)

    用于首次导入初始化账户结构 (侧边栏菜单)。
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    subhead_row_idx, _ = _find_subhead_row(rows)
    if subhead_row_idx is None:
        return []
    subhead_row = rows[subhead_row_idx]
    group_row = rows[subhead_row_idx - 1] if subhead_row_idx > 0 else []
    group_by_col = _forward_fill_group(group_row)

    out: list[BalanceSubhead] = []
    seen: set[tuple] = set()
    for ci, v in enumerate(subhead_row):
        if v in (None, ""):
            continue
        subhead = str(v).strip()
        if subhead in _IGNORED_SUBHEADS or len(subhead) > 16:
            continue
        group = GROUP_OVERRIDE.get(subhead) or group_by_col.get(ci, "")
        acc = _resolve_subhead(group, subhead)
        if not acc or not group:
            continue
        typ, owner, name = acc
        key = (group, typ, owner, name)
        if key in seen:
            continue
        seen.add(key)
        out.append(BalanceSubhead(
            group=group, type=typ, owner=owner, name=name,
        ))
    return out


@dataclass
class EntrySubhead:
    """年度账单表的小项结构 (大类=收入/支出)"""
    group: str       # 收入 / 支出
    name: str        # 小项 (君公司 / 物业费 ...)
    item_key: tuple  # (category, owner, sub_category, name)


def extract_entry_structure(ws, sheet_name: str) -> list[EntrySubhead]:
    """从年度账单表头提取 大类(收入/支出) -> 小项 结构

    遍历前 6 行表头, 凡命中 ITEM_HEADER_MAP 的列名即作为一个小项,
    大类取 item_key[0] (收入/支出)。不同年度表列名略有差异, 各表独立登记。
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    out: list[EntrySubhead] = []
    seen: set[tuple] = set()
    # 收集 (列号 -> 小项名) 跨行去重 (首次命中为准, 保留列顺序)
    col_name: dict[int, str] = {}
    for row in rows[:6]:
        for ci, v in enumerate(row):
            if v in (None, ""):
                continue
            t = str(v).strip()
            if t in ITEM_HEADER_MAP and ci not in col_name:
                col_name[ci] = t
    for ci in sorted(col_name):
        subhead = col_name[ci]
        item_key = ITEM_HEADER_MAP[subhead]
        group = item_key[0]  # 收入 / 支出
        key = (group, subhead, item_key)
        if key in seen:
            continue
        seen.add(key)
        out.append(EntrySubhead(group=group, name=subhead, item_key=item_key))
    return out


def _forward_fill_group(group_row) -> dict[int, str]:
    """对大类行做前向填充: 合并单元格仅在首列有值, 后续列为空 -> 继承左侧大类"""
    out: dict[int, str] = {}
    last = ""
    for ci, v in enumerate(group_row):
        if v not in (None, ""):
            text = str(v).strip()
            if text and len(text) <= 16:
                last = text
        if last:
            out[ci] = last
    return out


def parse_balance_sheet(ws, sheet_name: str) -> ParseResult:
    rows = list(ws.iter_rows(values_only=True))
    result = ParseResult(sheet_name=sheet_name, kind="balances")
    if not rows:
        return result

    subhead_row_idx, date_col = _find_subhead_row(rows)
    if subhead_row_idx is None:
        return result
    subhead_row = rows[subhead_row_idx]
    # 大类行 = 小项行上一行 (家庭结余: 小项在 row5, 大类合并单元格在 row4)
    group_row = rows[subhead_row_idx - 1] if subhead_row_idx > 0 else []
    group_by_col = _forward_fill_group(group_row)

    # 收集 (列号 -> (account_key, group))
    balance_cols: dict[int, tuple] = {}
    balance_groups: dict[int, str] = {}
    for ci, v in enumerate(subhead_row):
        if v in (None, ""):
            continue
        subhead = str(v).strip()
        if subhead in _IGNORED_SUBHEADS or len(subhead) > 16:
            continue
        group = GROUP_OVERRIDE.get(subhead) or group_by_col.get(ci, "")
        acc_key = _resolve_subhead(group, subhead)
        if not acc_key or not group:
            continue
        balance_cols[ci] = acc_key
        balance_groups[ci] = group

    years_seen = set()
    if date_col is None or not balance_cols:
        return result

    for row in rows[subhead_row_idx + 1:]:
        if not row or len(row) <= date_col:
            continue
        ym = _to_year_month(row[date_col])
        if ym is None:
            continue
        year, month = ym
        years_seen.add(year)

        for ci, acc_key in balance_cols.items():
            if ci >= len(row):
                continue
            val = _to_float(row[ci])
            if val is None:
                continue
            result.balances.append(ParsedBalance(
                year=year, month=month, account_key=acc_key, value=val,
                group=balance_groups[ci], sheet=sheet_name,
            ))
            result.rows_scanned += 1

    if years_seen:
        result.year_from = result.year_to = min(years_seen)
        if len(years_seen) > 1:
            result.year_to = max(years_seen)
    return result


# -------------------------------------------------------------- 主入口
def parse_workbook(path: str) -> list[ParseResult]:
    """解析整个 Excel 工作簿, 返回每个工作表的解析结果"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    results: list[ParseResult] = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        # 判断工作表类型
        if "结余" in sname or "盘点" in sname:
            kind = "balances"
        elif "年度账单" in sname or sname.startswith("君军之家"):
            kind = "entries"
        else:
            continue
        try:
            if kind == "entries":
                results.append(parse_year_sheet(ws, sname))
            else:
                results.append(parse_balance_sheet(ws, sname))
        except Exception as e:  # noqa: BLE001
            results.append(ParseResult(
                sheet_name=sname, kind="entries",
                rows_scanned=0,
            ))
            # 记录错误信息到一个特殊 entry note (避免崩溃)
            print(f"[excel_parser] parse {sname} error: {e}")
    wb.close()
    return results
