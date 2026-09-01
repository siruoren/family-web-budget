"""首次导入 Excel 初始化数据与结构

职责:
  1. 把 Excel 每个 sheet 页登记到 Sheet 表 (驱动左侧大菜单)
  2. 用家庭结余表头结构回填/补建 Account 的 sheet + group (大类)
  3. 给年度账单表的 Item 回填 sheet (group 即 category)
  4. 用 SheetColumn 登记每个 sheet 的 大类->小项 列结构 (驱动多级子菜单)
全部幂等: 重复执行不会重复创建, 只补齐缺失。
"""
import openpyxl
from sqlalchemy import select, func

from .. import db
from ..models import Sheet, Account, Item, SheetColumn
from .excel_parser import (
    parse_sheet_inventory, extract_balance_structure,
    extract_entry_structure,
)


def _ensure_sheet_records(path: str) -> int:
    """登记全部 sheet 页 (按 Excel 顺序), 返回新增数量"""
    infos = parse_sheet_inventory(path)
    existing = {
        s.name for s in db.session.execute(select(Sheet)).scalars()
    }
    added = 0
    for info in infos:
        if info.name in existing:
            # 更新 kind / sort_order (Excel 可能调整)
            row = db.session.execute(
                select(Sheet).where(Sheet.name == info.name)
            ).scalars().first()
            if row:
                row.kind = info.kind
                row.sort_order = info.order
            continue
        db.session.add(Sheet(
            name=info.name, kind=info.kind, sort_order=info.order,
            is_active=True, source_file=path.split("/")[-1],
        ))
        added += 1
    db.session.commit()
    return added


def _ensure_balance_accounts(path: str, sheet_name: str) -> tuple[int, int]:
    """用家庭结余结构补建账户并回填 sheet/group; 返回 (新建, 回填)"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return 0, 0
    ws = wb[sheet_name]
    subs = extract_balance_structure(ws, sheet_name)
    wb.close()

    # 已有账户缓存: (type, owner, name) -> Account
    accs = db.session.execute(select(Account)).scalars().all()
    cache: dict[tuple, Account] = {
        (a.type, a.owner, a.name): a for a in accs
    }

    created = 0
    backfilled = 0
    order = 0
    for s in subs:
        order += 1
        key = (s.type, s.owner, s.name)
        acc = cache.get(key)
        if acc is None:
            acc = Account(
                type=s.type, owner=s.owner, name=s.name,
                group=s.group, sheet=sheet_name,
                sort_order=order, is_active=True,
            )
            db.session.add(acc)
            db.session.flush()
            cache[key] = acc
            created += 1
        else:
            changed = False
            if not acc.group:
                acc.group = s.group
                changed = True
            if not acc.sheet:
                acc.sheet = sheet_name
                changed = True
            if changed:
                backfilled += 1
    db.session.commit()
    return created, backfilled


def _backfill_item_sheet(path: str) -> int:
    """年度账单条目: group 即 category (已存在), 这里回填 sheet 字段

    因条目跨多个年度表共享, sheet 统一记为来源工作表名中的最新一个
    (君军之家年度账单2026年始~); 已有 sheet 不覆盖。
    """
    infos = parse_sheet_inventory(path)
    entry_sheet = next(
        (i.name for i in infos if i.kind == "entries"), None
    )
    if not entry_sheet:
        return 0
    updated = 0
    for it in db.session.execute(
        select(Item).where(
            (Item.sheet == None) | (Item.sheet == "")  # noqa: E711
        )
    ).scalars():
        it.sheet = entry_sheet
        updated += 1
    db.session.commit()
    return updated


def _ensure_sheet_columns(path: str) -> int:
    """登记每个 sheet 的 大类->小项 列结构到 SheetColumn (幂等)

    - balances 类: 用 extract_balance_structure (大类=银行理财/...)
    - entries 类 : 用 extract_entry_structure (大类=收入/支出)
    返回新增列数。
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    infos = parse_sheet_inventory(path)
    existing = {
        (c.sheet_name, c.group, c.name)
        for c in db.session.execute(select(SheetColumn)).scalars()
    }
    added = 0
    for info in infos:
        if info.kind not in ("balances", "entries"):
            continue
        if info.name not in wb.sheetnames:
            continue
        ws = wb[info.name]
        if info.kind == "balances":
            subs = extract_balance_structure(ws, info.name)
            rows = [
                (s.group, s.name, f"{s.type}|{s.owner}|{s.name}")
                for s in subs
            ]
        else:
            subs = extract_entry_structure(ws, info.name)
            rows = [
                (s.group, s.name, "|".join(str(x) for x in s.item_key))
                for s in subs
            ]
        for order, (group, name, key) in enumerate(rows):
            if (info.name, group, name) in existing:
                continue
            db.session.add(SheetColumn(
                sheet_name=info.name, group=group, name=name,
                item_key=key, sort_order=order,
            ))
            existing.add((info.name, group, name))
            added += 1
    db.session.commit()
    wb.close()
    return added


def initialize_structure_from_excel(path: str) -> dict:
    """首次/手动初始化: 登记 sheet + 补建账户结构 + 回填条目 sheet + 列结构"""
    summary = {"sheets_added": 0, "accounts_created": 0,
               "accounts_backfilled": 0, "items_backfilled": 0,
               "columns_added": 0}
    summary["sheets_added"] = _ensure_sheet_records(path)
    # 找结余 sheet (家庭结余)
    infos = parse_sheet_inventory(path)
    bal_sheet = next(
        (i.name for i in infos if i.kind == "balances"), None
    )
    if bal_sheet:
        c, b = _ensure_balance_accounts(path, bal_sheet)
        summary["accounts_created"] = c
        summary["accounts_backfilled"] = b
    summary["items_backfilled"] = _backfill_item_sheet(path)
    summary["columns_added"] = _ensure_sheet_columns(path)
    return summary
