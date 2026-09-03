"""家庭统计表 -> 系统可导入历史数据 CSV 转换脚本

读取 统计表2021年~20260501.xlsx 中 6 个年度账单 sheet (2021-2026),
按统一的列映射规则, 将每月各收支科目规范化为长表 CSV:

    年份,月份,类型,属主,条目名称,金额,备注

输出: exports/family_history_import.csv (UTF-8-SIG, 便于 Excel 打开)

设计要点:
- 跨年度统一条目名称 (如 "君工资" 在 2021-2026 都用同一名称, 导入后关联同一 AccountItem)
- 跳过所有合计/小计列 (收入合计/支出合计/月总收入/月总支出 等)
- 跳过 "月增长/月增减" (变化量, 非结余, 避免与结余重复)
- 空值单元格跳过 (该月该科目未发生 -> 不生成 Asset 记录)
- 金额保留负数 (如理财亏损), 四舍五入到 2 位
"""
import csv
import os
from datetime import datetime

import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_PATH = os.path.join(BASE_DIR, "统计表2021年~20260501.xlsx")
OUT_PATH = os.path.join(BASE_DIR, "exports", "family_history_import.csv")

# 各年度 sheet 的列映射: (列号, 条目名称, 类型, 属主)
# 列号对应 openpyxl 的 1-based 列索引 (B=2, C=3, ...)
# 未列入的列 = 合计/小计/变化量, 自动跳过
SHEET_MAP = {
    "君军之家2021年": [
        # 收入 (按军/君分组)
        (3, "军工资", "收入", "军"),
        (4, "军其他收入", "收入", "军"),
        (6, "君工资", "收入", "君"),
        (7, "君学校收入", "收入", "君"),
        (8, "君其他收入", "收入", "君"),
        # 支出 (固定支出 + 生活消费)
        (10, "房租", "支出", "家庭"),
        (11, "公积金贷款", "支出", "家庭"),
        (12, "商业贷款", "支出", "家庭"),
        (20, "生活消费", "支出", "家庭"),
        # 结余 (各人月末余额)
        (14, "军月末结余", "结余", "军"),
        (16, "君月末结余", "结余", "君"),
    ],
    "君军之家2022年": [
        (3, "君工资", "收入", "君"),
        (4, "君其他收入", "收入", "君"),
        (5, "军工资", "收入", "军"),
        (6, "军其他收入", "收入", "军"),
        (7, "其他收入", "收入", "家庭"),
        (9, "公积金贷款", "支出", "家庭"),
        (10, "商业贷款", "支出", "家庭"),
        (11, "车贷", "支出", "家庭"),
        (12, "房租", "支出", "家庭"),
        (13, "水费", "支出", "家庭"),
        (14, "电费", "支出", "家庭"),
        (15, "燃气费", "支出", "家庭"),
        (17, "其他支出", "支出", "家庭"),
        (18, "生活支出", "支出", "家庭"),
        (20, "月末结余", "结余", "家庭"),
    ],
    "君军之家2023年": [
        (3, "君工资", "收入", "君"),
        (4, "君其他收入", "收入", "君"),
        (5, "军工资", "收入", "军"),
        (6, "军其他收入", "收入", "军"),
        (7, "理财收益", "收入", "家庭"),
        (8, "其他收入", "收入", "家庭"),
        (10, "公积金贷款", "支出", "家庭"),
        (11, "房租", "支出", "家庭"),
        (12, "水费", "支出", "家庭"),
        (13, "电费", "支出", "家庭"),
        (14, "燃气费", "支出", "家庭"),
        (16, "其他支出", "支出", "家庭"),
        (17, "生活支出", "支出", "家庭"),
        (19, "月末结余", "结余", "家庭"),
    ],
    "君军之家2024年 ": [
        (3, "君工资", "收入", "君"),
        (4, "君其他收入", "收入", "君"),
        (5, "军工资", "收入", "军"),
        (6, "军其他收入", "收入", "军"),
        (7, "理财收益", "收入", "家庭"),
        (8, "其他收入", "收入", "家庭"),
        (10, "公积金贷款", "支出", "家庭"),
        (11, "物业费", "支出", "家庭"),
        (12, "房租", "支出", "家庭"),
        (13, "水费", "支出", "家庭"),
        (14, "电费", "支出", "家庭"),
        (15, "燃气费", "支出", "家庭"),
        (17, "其他支出", "支出", "家庭"),
        (19, "月末结余", "结余", "家庭"),
    ],
    "君军之家2025年 ": [
        (3, "君工资", "收入", "君"),
        (4, "君其他收入", "收入", "君"),
        (5, "军工资", "收入", "军"),
        (6, "军其他收入", "收入", "军"),
        (7, "理财收益", "收入", "家庭"),
        (8, "其他收入", "收入", "家庭"),
        (10, "物业费", "支出", "家庭"),
        (11, "房租", "支出", "家庭"),
        (12, "水费", "支出", "家庭"),
        (13, "电费", "支出", "家庭"),
        (14, "燃气费", "支出", "家庭"),
        (16, "其他支出", "支出", "家庭"),
        (18, "月末结余", "结余", "家庭"),
    ],
    "君军之家年度账单2026年始~": [
        (3, "君工资", "收入", "君"),
        (4, "君其他收入", "收入", "君"),
        (5, "军工资", "收入", "军"),
        (6, "公积金提取", "收入", "家庭"),
        (7, "理财收益", "收入", "家庭"),
        (8, "其他收入", "收入", "家庭"),
        (10, "信用及现金支出", "支出", "家庭"),
        (11, "差额支出", "支出", "家庭"),
        (13, "月末结余", "结余", "家庭"),
    ],
}

# 年份识别: 从 sheet 名取年份数字, 用于备注追溯
SHEET_YEAR = {
    "君军之家2021年": 2021,
    "君军之家2022年": 2022,
    "君军之家2023年": 2023,
    "君军之家2024年 ": 2024,
    "君军之家2025年 ": 2025,
    "君军之家年度账单2026年始~": 2026,  # 含2026及之后规划数据
}


def to_float(v):
    """单元格值转 float; 无法解析返回 None (含空值)"""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    s = str(v).strip().replace(",", "").replace("，", "")
    if not s or s in ("-", "—", "——"):
        return None
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def main():
    if not os.path.exists(XLSX_PATH):
        raise SystemExit(f"找不到 Excel: {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)

    records = []          # (year, month, type, owner, name, value, note)
    per_sheet_count = {}
    skipped_rows = 0

    for sheet_name, col_map in SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"[警告] 缺少 sheet: {sheet_name!r}, 跳过")
            continue
        ws = wb[sheet_name]
        cnt = 0

        # 找 B 列为日期的行 (数据行), 跳过标题/期初余额
        for r in range(5, ws.max_row + 1):
            dv = ws.cell(r, 2).value
            if not isinstance(dv, datetime):
                continue
            year, month = dv.year, dv.month
            for col, name, itype, owner in col_map:
                val = to_float(ws.cell(r, col).value)
                if val is None:
                    continue  # 空值: 该月该科目未发生, 不生成记录
                note = f"源自{sheet_name.strip()}"
                records.append((year, month, itype, owner, name, val, note))
                cnt += 1
        per_sheet_count[sheet_name.strip()] = cnt

    # 写 CSV (UTF-8-SIG)
    with open(OUT_PATH, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["年份", "月份", "类型", "属主", "条目名称", "金额", "备注"])
        for rec in records:
            w.writerow(rec)

    # ---- 统计摘要 ----
    print("=" * 60)
    print(f"转换完成: 共 {len(records)} 条月度记录")
    print(f"输出: {OUT_PATH}")
    print("-" * 60)
    print("各 sheet 记录数:")
    for k, v in per_sheet_count.items():
        print(f"  {k}: {v} 条")

    # 去重条目清单
    items = {}
    for y, m, t, o, n, v, note in records:
        key = (n, t, o)
        items.setdefault(key, {"count": 0, "sum": 0.0, "months": set()})
        items[key]["count"] += 1
        items[key]["sum"] += v
        items[key]["months"].add((y, m))

    print("-" * 60)
    print(f"去重条目数: {len(items)} 个")
    print(f"{'条目名称':<16}{'类型':<6}{'属主':<6}{'记录数':<6}{'月份覆盖':<8}{'合计金额':>14}")
    # 按类型+属主+名称排序
    for (n, t, o), info in sorted(items.items(), key=lambda x: (x[0][1], x[0][2], x[0][0])):
        print(f"{n:<16}{t:<6}{o:<6}{info['count']:<6}{len(info['months']):<8}{info['sum']:>14.2f}")

    # 年份覆盖
    years = sorted({r[0] for r in records})
    print("-" * 60)
    print(f"年份覆盖: {years}")
    print(f"总月份: {len({(r[0], r[1]) for r in records})} 个")


if __name__ == "__main__":
    main()
